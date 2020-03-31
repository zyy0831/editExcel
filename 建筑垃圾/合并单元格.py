# ⑤ 合并和拆分单元格
# https://www.cnblogs.com/programmer-tlh/p/10461353.html
# 所谓合并单元格，即以合并区域的左上角的那个单元格为基准，覆盖其他单元格使之称为一个大的单元格。
# 相反，拆分单元格后将这个大单元格的值返回到原来的左上角位置。
# # 合并单元格， 往左上角写入数据即可
# sheet.merge_cells('B1:G1') # 合并一行中的几个单元格
# sheet.merge_cells('A1:C3') # 合并一个矩形区域中的单元格
# 合并后只可以往左上角写入数据，也就是区间中:左边的坐标。
# 如果这些要合并的单元格都有数据，只会保留左上角的数据，其他则丢弃。换句话说若合并前不是在左上角写入数据，合并后单元格中不会有数据。
# 以下是拆分单元格的代码。拆分后，值回到A1位置
# sheet.unmerge_cells('A1:C3')

from openpyxl  import load_workbook
import os

data_path = r'C:\Users\zyy\Desktop\1'
for fileName in os.listdir(data_path):
    fileName_Path = os.path.join(data_path, fileName)#拼接路径
    wb = load_workbook(fileName_Path)
    ws = wb.active
    rows=ws.max_row
    columndata=[]
    for i in range(1,rows+1):
        cellvalue = ws.cell(row=i,column=1).value
        cellvalue2 = ws.cell(row=i+1,column=1).value
        if cellvalue == cellvalue2:
            columndata.append(i)
        else:
            if len(columndata)>1:
                a='A'+str(min(columndata))
                b='A'+str(max(columndata))
                C=a+':'+b
                ws.merge_cells(C)
                columndata=[]
    wb.save(fileName_Path)
    wb.close()
#https://www.jianshu.com/p/42f1d2909bb6
#https://www.jianshu.com/p/2e35f62d9d5d

